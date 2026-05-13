import csv
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from core.models import CashEntry, Client, Contract, Invoice
from core.serializers import competence_label


def _money_br(value: Decimal | float) -> str:
    v = float(value)
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def queryset_for_month(year: int, month: int) -> tuple[list[CashEntry], list[Invoice]]:
    cash = list(
        CashEntry.objects.filter(due_date__year=year, due_date__month=month)
        .select_related("contract")
        .order_by("due_date", "id")
    )
    invoices = list(
        Invoice.objects.filter(competence__year=year, competence__month=month)
        .select_related("client")
        .order_by("id")
    )
    return cash, invoices


def monthly_summary(cash: list[CashEntry]) -> dict:
    income = sum((e.amount for e in cash if e.kind == CashEntry.Kind.ENTRADA), Decimal("0"))
    expense = sum((e.amount for e in cash if e.kind == CashEntry.Kind.SAIDA), Decimal("0"))
    return {
        "receitas": income,
        "despesas": expense,
        "resultado": income - expense,
    }


def build_monthly_csv(year: int, month: int) -> str:
    cash, invoices = queryset_for_month(year, month)
    summary = monthly_summary(cash)
    comp = date(year, month, 1)
    label = competence_label(comp)

    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(["DataGreen — exportação mensal para contador"])
    w.writerow(["Competência", label])
    w.writerow([])
    w.writerow(["Resumo"])
    w.writerow(["Receitas (lançamentos de entrada no mês)", _money_br(summary["receitas"])])
    w.writerow(["Despesas (lançamentos de saída no mês)", _money_br(summary["despesas"])])
    w.writerow(["Resultado", _money_br(summary["resultado"])])
    w.writerow([])
    w.writerow(["Lançamentos financeiros", "Tipo", "Categoria", "Valor", "Vencimento", "Status", "Contrato"])
    for e in cash:
        contract_title = e.contract.title if e.contract else ""
        w.writerow(
            [
                e.description,
                e.kind,
                e.category,
                _money_br(e.amount),
                e.due_date.isoformat(),
                e.status,
                contract_title,
            ]
        )
    w.writerow([])
    w.writerow(["Notas fiscais", "Cliente", "Valor", "Status"])
    for inv in invoices:
        w.writerow([inv.number, inv.client.name, _money_br(inv.amount), inv.status])
    w.writerow([])
    w.writerow(["Clientes ativos (cadastro)", str(Client.objects.filter(active=True).count())])
    return buf.getvalue()


def csv_response(year: int, month: int) -> HttpResponse:
    content = build_monthly_csv(year, month)
    resp = HttpResponse(content.encode("utf-8-sig"), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="datagreen_{year}_{month:02d}.csv"'
    return resp


def xlsx_response(year: int, month: int) -> HttpResponse:
    cash, invoices = queryset_for_month(year, month)
    summary = monthly_summary(cash)
    comp = date(year, month, 1)
    label = competence_label(comp)

    wb = Workbook()
    bold = Font(bold=True)

    ws0 = wb.active
    ws0.title = "Resumo"
    ws0.append(["DataGreen — exportação mensal"])
    ws0.append(["Competência", label])
    ws0.append([])
    ws0.append(["Receitas (entradas)", float(summary["receitas"])])
    ws0.append(["Despesas (saídas)", float(summary["despesas"])])
    ws0.append(["Resultado", float(summary["resultado"])])
    ws0.append(["Clientes ativos", Client.objects.filter(active=True).count()])
    for row in ws0.iter_rows(min_row=4, max_row=6, min_col=1, max_col=2):
        row[0].font = bold

    ws1 = wb.create_sheet("Lançamentos")
    ws1.append(["Descrição", "Tipo", "Categoria", "Valor", "Vencimento", "Status", "Contrato"])
    for cell in ws1[1]:
        cell.font = bold
    for e in cash:
        ws1.append(
            [
                e.description,
                e.kind,
                e.category,
                float(e.amount),
                e.due_date.isoformat(),
                e.status,
                e.contract.title if e.contract else "",
            ]
        )

    ws2 = wb.create_sheet("Notas")
    ws2.append(["Número", "Cliente", "Valor", "Status"])
    for cell in ws2[1]:
        cell.font = bold
    for inv in invoices:
        ws2.append([inv.number, inv.client.name, float(inv.amount), inv.status])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="datagreen_{year}_{month:02d}.xlsx"'
    return resp


def full_xlsx_response() -> HttpResponse:
    wb = Workbook()
    bold = Font(bold=True)

    ws_clients = wb.active
    ws_clients.title = "Clientes"
    ws_clients.append(["ID", "Nome", "Tipo", "Telefone", "Cidade", "Ativo"])
    for cell in ws_clients[1]:
        cell.font = bold
    for client in Client.objects.order_by("name"):
        ws_clients.append([client.id, client.name, client.type, client.phone, client.city, "Sim" if client.active else "Não"])

    ws_contracts = wb.create_sheet("Contratos")
    ws_contracts.append(["ID", "Cliente", "Título", "Tipo", "Valor mensal", "Status", "Início", "Fim"])
    for cell in ws_contracts[1]:
        cell.font = bold
    for contract in Contract.objects.select_related("client").order_by("-id"):
        ws_contracts.append(
            [
                contract.id,
                contract.client.name,
                contract.title,
                contract.kind,
                float(contract.monthly_value),
                contract.status,
                contract.start_date.isoformat() if contract.start_date else "",
                contract.end_date.isoformat() if contract.end_date else "",
            ]
        )

    ws_cash = wb.create_sheet("Financeiro")
    ws_cash.append(["ID", "Tipo", "Descrição", "Categoria", "Valor", "Vencimento", "Status", "Contrato"])
    for cell in ws_cash[1]:
        cell.font = bold
    for entry in CashEntry.objects.select_related("contract").order_by("-due_date", "-id"):
        ws_cash.append(
            [
                entry.id,
                entry.kind,
                entry.description,
                entry.category,
                float(entry.amount),
                entry.due_date.isoformat(),
                entry.status,
                entry.contract.title if entry.contract else "",
            ]
        )

    ws_invoices = wb.create_sheet("Notas fiscais")
    ws_invoices.append(["ID", "Cliente", "Número", "Valor", "Status", "Competência", "Emissão", "Vencimento"])
    for cell in ws_invoices[1]:
        cell.font = bold
    for inv in Invoice.objects.select_related("client").order_by("-competence", "-id"):
        ws_invoices.append(
            [
                inv.id,
                inv.client.name,
                inv.number,
                float(inv.amount),
                inv.status,
                inv.competence.isoformat(),
                inv.issued_at.isoformat() if inv.issued_at else "",
                inv.due_date.isoformat() if inv.due_date else "",
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="datagreen_full_export.xlsx"'
    return resp
